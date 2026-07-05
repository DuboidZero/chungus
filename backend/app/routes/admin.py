from app.models.cohort import Cohort
from io import BytesIO
from datetime import datetime
from app.schemas.academic import SemesterCreate, SemesterUpdate, SemesterResponse
from app.core.grading import calculate_sgpa
from app.models.mentor import MentorAssignment
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, status
from sqlalchemy.orm import Session
from sqlalchemy import or_
from sqlalchemy.exc import SQLAlchemyError
from openpyxl import load_workbook


from app.models.user import User
from app.models.academic import Semester, Subject
from app.models.academic_structure import Branch, Course
from app.schemas.admin import (
    BulkUploadResult, SeededStudent, TeacherCreate, UserSummary,
    MentorAssignmentRequest, MentorAssignmentResponse,
    AdminUserUpdate, AdminUserResponse, ResetPasswordResponse, ToggleStatusResponse,
    CohortMentorUpdate, SeededTeacher, BulkTeacherUploadResult,
)
from app.core.dependencies import get_current_admin, get_db
from app.core.security import hash_password, generate_temp_password

router = APIRouter(prefix="/admin", tags=["admin"])


# ============================================================
#  Shared import helpers
# ============================================================
def _open_sheet(contents: bytes):
    try:
        wb = load_workbook(BytesIO(contents))
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid Excel file")
    return wb.active


def _student_by_prn(db: Session, prn) -> User | None:
    if prn is None:
        return None
    return db.query(User).filter(User.prn == str(prn).strip(), User.role == "student").first()


def _parse_date(v):
    """Accept a date, 'YYYY-MM-DD', or 'YYYY-MM'. Return a date or None."""
    if v is None or v == "":
        return None
    if hasattr(v, "date"):        # datetime -> date
        return v.date()
    if isinstance(v, str):
        s = v.strip()
        if len(s) == 7 and s[4] == "-":   # YYYY-MM
            s = s + "-01"
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except ValueError:
            raise ValueError(f"invalid date '{v}' (use YYYY-MM-DD)")
    raise ValueError(f"invalid date '{v}'")

def _find_or_create_cohort(db: Session, academic_year: str, department: str) -> Cohort:
    """Return the cohort for this (year, dept), creating it if it doesn't exist."""
    cohort = db.query(Cohort).filter(
        Cohort.academic_year == academic_year,
        Cohort.department == department,
    ).first()
    if cohort is None:
        cohort = Cohort(academic_year=academic_year, department=department)
        db.add(cohort)
        db.flush()
    return cohort


# ============================================================
#  Bulk student account upload
#  Columns: A=PRN  B=Full Name  C=Year of Birth  D=Batch  E=Branch
# ============================================================
@router.post("/import/students", response_model=BulkUploadResult)
async def bulk_upload_students(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    sheet = _open_sheet(await file.read())

    created = 0
    skipped = 0
    errors = []
    seeded_students = []
    seen_prns = set()   # track PRNs within THIS upload to catch in-file duplicates

    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(cell is None for cell in row):
            continue

        prn, full_name, year_of_birth, batch, branch, academic_year = (list(row) + [None] * 6)[:6]

        if not prn or not full_name or not year_of_birth:
            errors.append(f"Row {row_num}: missing PRN, name, or year of birth")
            continue

        prn = str(prn).strip()

        # Skip if PRN already in DB OR already seen earlier in this same file
        if prn in seen_prns or db.query(User).filter(User.prn == prn).first():
            skipped += 1
            continue

        try:
            year = int(year_of_birth)
            dept = str(branch).strip() if branch else None
            ac_year = str(academic_year).strip() if academic_year else None
            # Random temp password (returned to the admin for private distribution).
            # Name+YOB passwords are guessable, so anyone could hijack an account
            # before the student's first login.
            initial_password = generate_temp_password()
            user = User(
                role="student",
                prn=prn,
                name=str(full_name),
                department=dept,
                academic_year=ac_year,
                batch=str(batch).strip() if batch else None,
                hashed_password=hash_password(initial_password),
                must_change_password=True,
            )
            db.add(user)
            db.flush()  # surface any DB error now, on this row

            # Auto-create the cohort for this (academic_year + department) if needed
            if ac_year and dept:
                _find_or_create_cohort(db, ac_year, dept)

            seen_prns.add(prn)
            created += 1
            seeded_students.append(SeededStudent(
                prn=prn,
                full_name=str(full_name),
                initial_password=initial_password,
            ))
        except (ValueError, TypeError, SQLAlchemyError) as e:
            db.rollback()
            errors.append(f"Row {row_num}: {str(e)}")

    db.commit()

    return BulkUploadResult(
        created=created,
        skipped=skipped,
        errors=errors,
        students=seeded_students,
    )

@router.get("/cohorts")
async def list_cohorts(db: Session = Depends(get_db), admin: User = Depends(get_current_admin)):
    """List all cohorts with student count and assigned mentor."""
    cohorts = db.query(Cohort).all()
    result = []
    for c in cohorts:
        # students in this cohort = matching year + department
        student_count = db.query(User).filter(
            User.role == "student",
            User.academic_year == c.academic_year,
            User.department == c.department,
        ).count()
        mentor = db.query(User).filter(User.id == c.mentor_id).first() if c.mentor_id else None
        result.append({
            "id": c.id,
            "academicYear": c.academic_year,
            "department": c.department,
            "studentCount": student_count,
            "academicMentorId": c.mentor_id,
            "academicMentorName": mentor.name if mentor else None,
        })
    return result

# ============================================================
#  Bulk teacher account upload
#  Columns: A=Email  B=Full Name  C=Department
# ============================================================
@router.post("/import/teachers", response_model=BulkTeacherUploadResult)
async def bulk_upload_teachers(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    sheet = _open_sheet(await file.read())

    created = 0
    skipped = 0
    errors = []
    seeded_teachers = []
    seen_emails = set()   # track emails within THIS upload to catch in-file duplicates

    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(cell is None for cell in row):
            continue

        email, full_name, department = (list(row) + [None] * 3)[:3]

        if not email or not full_name:
            errors.append(f"Row {row_num}: missing email or full name")
            continue

        email = str(email).strip().lower()

        if "@" not in email or "." not in email.split("@")[-1]:
            errors.append(f"Row {row_num}: invalid email '{email}'")
            continue

        # Skip if email already in DB OR already seen earlier in this same file
        if email in seen_emails or db.query(User).filter(User.email == email).first():
            skipped += 1
            continue

        try:
            initial_password = generate_temp_password()
            user = User(
                role="teacher",
                email=email,
                name=str(full_name).strip(),
                department=str(department).strip() if department else None,
                hashed_password=hash_password(initial_password),
                must_change_password=True,
            )
            db.add(user)
            db.flush()  # surface any DB error now, on this row

            seen_emails.add(email)
            created += 1
            seeded_teachers.append(SeededTeacher(
                email=email,
                full_name=str(full_name).strip(),
                initial_password=initial_password,
            ))
        except (ValueError, TypeError, SQLAlchemyError) as e:
            db.rollback()
            errors.append(f"Row {row_num}: {str(e)}")

    db.commit()

    return BulkTeacherUploadResult(
        created=created,
        skipped=skipped,
        errors=errors,
        teachers=seeded_teachers,
    )

# ============================================================
#  Bulk course/subject upload (per branch)
#  Columns: A=Semester  B=Course Code  C=Course Name  D=Type  E=Credits  F=Marking Scheme
# ============================================================
@router.post("/import/courses", response_model=dict)
async def bulk_upload_courses(
    branch_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    branch = db.query(Branch).filter(Branch.id == branch_id).first()
    if branch is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Branch not found")

    sheet = _open_sheet(await file.read())

    created = 0
    skipped = 0
    errors = []
    seen_codes = set()   # track course codes within THIS upload to catch in-file duplicates

    valid_types = {"DSC", "SEC", "VEC", "AEC"}

    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(cell is None for cell in row):
            continue

        semester, course_code, course_name, course_type, credits, marking_scheme = (list(row) + [None] * 6)[:6]

        if not semester or not course_code or not course_name or not course_type or credits is None:
            errors.append(f"Row {row_num}: missing semester, course code, name, type, or credits")
            continue

        course_code = str(course_code).strip()

        if course_code in seen_codes or db.query(Course).filter(
            Course.branch_id == branch_id, Course.course_code == course_code
        ).first():
            skipped += 1
            continue

        try:
            sem_int = int(semester)
            if not (1 <= sem_int <= 6):
                raise ValueError(f"semester must be 1-6, got {sem_int}")

            course_type = str(course_type).strip().upper()
            if course_type not in valid_types:
                raise ValueError(f"type must be one of {sorted(valid_types)}, got '{course_type}'")

            credits_int = int(credits)

            course = Course(
                branch_id=branch_id,
                semester=sem_int,
                course_code=course_code,
                course_name=str(course_name).strip(),
                type=course_type,
                credits=credits_int,
                marking_scheme=str(marking_scheme).strip() if marking_scheme else None,
            )
            db.add(course)
            db.flush()  # surface any DB error now, on this row

            seen_codes.add(course_code)
            created += 1
        except (ValueError, TypeError, SQLAlchemyError) as e:
            db.rollback()
            errors.append(f"Row {row_num}: {str(e)}")

    db.commit()

    return {"created": created, "skipped": skipped, "errors": errors}

# ============================================================
#  Bulk marks upload (one row per subject)
#  Columns: A=PRN  B=Semester  C=Subject  D=Obtained  E=Max  F=Credits
#  GPA/CGPA are computed from these marks by the grading code.
# ============================================================
@router.post("/import/academic-records", response_model=dict)
async def bulk_upload_marks(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    sheet = _open_sheet(await file.read())
    created = 0
    skipped = 0
    errors = []
    semester_cache = {}  # (user_id, semester_number) -> Semester

    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(cell is None for cell in row):
            continue

        prn, semester_number, subject_name, obtained, max_marks, credits = (list(row) + [None] * 6)[:6]

        if not prn or not semester_number or not subject_name:
            errors.append(f"Row {row_num}: missing PRN, semester, or subject")
            continue

        prn = str(prn).strip()
        student = db.query(User).filter(User.prn == prn, User.role == "student").first()
        if not student:
            errors.append(f"Row {row_num}: no student with PRN {prn}")
            continue

        try:
            sem_num = int(semester_number)
            cache_key = (student.id, sem_num)

            if cache_key in semester_cache:
                semester = semester_cache[cache_key]
            else:
                semester = db.query(Semester).filter(
                    Semester.user_id == student.id,
                    Semester.semester_number == sem_num,
                ).first()
                if not semester:
                    semester = Semester(user_id=student.id, semester_number=sem_num)
                    db.add(semester)
                    db.flush()
                semester_cache[cache_key] = semester

            db.add(Subject(
                semester_id=semester.id,
                name=str(subject_name),
                marks_obtained=float(obtained) if obtained is not None else None,
                max_marks=float(max_marks) if max_marks is not None else None,
                grade=None,  # derived from marks % by the grading calculator
                credits=int(credits) if credits else 0,
            ))
            db.flush()
            created += 1
        except (ValueError, TypeError, SQLAlchemyError) as e:
            db.rollback()
            semester_cache.clear()  # cached semester objects are invalid after rollback
            errors.append(f"Row {row_num}: {str(e)}")

    db.commit()
    # keep subjectsAdded for backwards-compat, add created for consistency with other imports
    return {"created": created, "subjectsAdded": created, "skipped": 0, "errors": errors}


# ============================================================
#  Teacher account creation
# ============================================================
@router.post("/teachers", response_model=UserSummary, status_code=status.HTTP_201_CREATED)
async def create_teacher(
    payload: TeacherCreate,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    if db.query(User).filter(User.email == payload.email).first():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Email already exists")
    teacher = User(
        role="teacher",
        email=payload.email,
        name=payload.full_name,
        hashed_password=hash_password(payload.password),
        must_change_password=True,
    )
    db.add(teacher)
    db.commit()
    db.refresh(teacher)
    return teacher


# ============================================================
#  Account management: list users + activate/deactivate
# ============================================================
@router.get("/users", response_model=list[UserSummary])
async def list_users(
    role: str | None = None,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    query = db.query(User)
    if role:
        query = query.filter(User.role == role)
    return query.all()


@router.patch("/users/{user_id}/active", response_model=UserSummary)
async def set_user_active(
    user_id: str,
    is_active: bool,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    user = db.query(User).filter(User.id == user_id).first()
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    user.is_active = is_active
    db.commit()
    db.refresh(user)
    return user


# @router.post("/assignments", response_model=MentorAssignmentResponse, status_code=status.HTTP_201_CREATED)
# async def assign_mentor(
#     payload: MentorAssignmentRequest,
#     db: Session = Depends(get_db),
#     admin: User = Depends(get_current_admin),
# ):
#     # validate both users exist and have correct roles
#     teacher = db.query(User).filter(User.id == payload.teacher_id, User.role == "teacher").first()
#     if not teacher:
#         raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Teacher not found")
#     student = db.query(User).filter(User.id == payload.student_id, User.role == "student").first()
#     if not student:
#         raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Student not found")

#     # one mentor per student: replace existing assignment if any
#     existing = db.query(MentorAssignment).filter(MentorAssignment.student_id == payload.student_id).first()
#     if existing:
#         existing.teacher_id = payload.teacher_id
#         db.commit()
#         db.refresh(existing)
#         return existing

#     assignment = MentorAssignment(teacher_id=payload.teacher_id, student_id=payload.student_id)
#     db.add(assignment)
#     db.commit()
#     db.refresh(assignment)
#     return assignment


# ============================================================
#  User management (per admin contract)
# ============================================================
@router.patch("/users/{user_id}", response_model=AdminUserResponse)
async def update_user(
    user_id: str,
    payload: AdminUserUpdate,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    user = db.query(User).filter(User.id == user_id).first()
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    updates = payload.model_dump(exclude_unset=True)

    # Uniqueness check for email before applying
    new_email = updates.get("email")
    if new_email and new_email != user.email:
        clash = db.query(User).filter(User.email == new_email, User.id != user_id).first()
        if clash:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Email already in use")

    # Only set columns that actually exist on the model (academic_year/batch come later)
    settable = {"name", "email", "department", "academic_year", "batch"}
    for field, value in updates.items():
        if field in settable and hasattr(user, field):
            setattr(user, field, value)

    db.commit()
    db.refresh(user)
    return user


@router.post("/users/{user_id}/reset-password", response_model=ResetPasswordResponse)
async def reset_user_password(
    user_id: str,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    user = db.query(User).filter(User.id == user_id).first()
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    # Generate a fresh temp password; force change on next login
    temp_password = generate_temp_password()
    user.hashed_password = hash_password(temp_password)
    user.must_change_password = True
    db.commit()

    return ResetPasswordResponse(id=user.id, temporary_password=temp_password, first_login=True)


@router.post("/users/{user_id}/toggle-status", response_model=ToggleStatusResponse)
async def toggle_user_status(
    user_id: str,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    user = db.query(User).filter(User.id == user_id).first()
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    # Best-practice guards: admin can't lock themselves out or remove the last admin
    if user.id == admin.id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="You cannot change your own status")
    if user.role == "admin" and user.is_active:
        active_admins = db.query(User).filter(User.role == "admin", User.is_active == True).count()
        if active_admins <= 1:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cannot deactivate the last active admin")

    user.is_active = not user.is_active
    db.commit()
    db.refresh(user)

    # contract: deactivated = true means the account is now OFF
    return ToggleStatusResponse(id=user.id, deactivated=not user.is_active)


# ============================================================
#  Admin: Academic Records CRUD (per admin contract)
#  POST/PATCH/DELETE /admin/students/:id/academic-records
# ============================================================

def _build_semester_response(semester: Semester) -> dict:
    """Shape a semester with computed gpa + total_credits + subjects."""
    subjects = semester.subjects
    return {
        "id": semester.id,
        "semester_number": semester.semester_number,
        "gpa": calculate_sgpa(subjects),
        "total_credits": sum(s.credits for s in subjects),
        "subjects": subjects,
        "created_at": semester.created_at,
        "updated_at": semester.updated_at,
    }


def _get_student_or_404(db: Session, student_id: str) -> User:
    """Fetch a user that must exist AND be a student."""
    student = db.query(User).filter(User.id == student_id, User.role == "student").first()
    if student is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Student not found")
    return student


@router.post("/students/{student_id}/academic-records", response_model=SemesterResponse, status_code=status.HTTP_201_CREATED)
async def admin_create_academic_record(
    student_id: str,
    payload: SemesterCreate,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    _get_student_or_404(db, student_id)

    # Prevent duplicate semester numbers for the same student
    existing = db.query(Semester).filter(
        Semester.user_id == student_id,
        Semester.semester_number == payload.semester_number,
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Semester {payload.semester_number} already exists for this student",
        )

    semester = Semester(user_id=student_id, semester_number=payload.semester_number)
    db.add(semester)
    db.flush()  # get semester.id before adding subjects
    for subj in payload.subjects:
        db.add(Subject(semester_id=semester.id, **subj.model_dump()))
    db.commit()
    db.refresh(semester)
    return _build_semester_response(semester)


@router.patch("/students/{student_id}/academic-records/{semester_id}", response_model=SemesterResponse)
async def admin_update_academic_record(
    student_id: str,
    semester_id: str,
    payload: SemesterUpdate,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    _get_student_or_404(db, student_id)

    semester = db.query(Semester).filter(
        Semester.id == semester_id,
        Semester.user_id == student_id,
    ).first()
    if semester is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Semester not found for this student")

    if payload.semester_number is not None:
        semester.semester_number = payload.semester_number

    # If subjects provided, replace the whole set
    if payload.subjects is not None:
        db.query(Subject).filter(Subject.semester_id == semester.id).delete()
        for subj in payload.subjects:
            db.add(Subject(semester_id=semester.id, **subj.model_dump()))

    db.commit()
    db.refresh(semester)
    return _build_semester_response(semester)


@router.delete("/students/{student_id}/academic-records/{semester_id}", status_code=status.HTTP_204_NO_CONTENT)
async def admin_delete_academic_record(
    student_id: str,
    semester_id: str,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    _get_student_or_404(db, student_id)

    semester = db.query(Semester).filter(
        Semester.id == semester_id,
        Semester.user_id == student_id,
    ).first()
    if semester is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Semester not found for this student")

    db.query(Subject).filter(Subject.semester_id == semester_id).delete()
    db.delete(semester)
    db.commit()


# ============================================================
#  Bulk imports — Skills, Projects, Achievements, Experience
#  Every Excel's first column is PRN (links row to a student).
#  Remaining columns mirror the manual-entry API fields.
#  All return {created, skipped, errors[]}.
#  Each row is flushed independently so one bad row does not
#  sink the whole upload (per-row resilience).
# ============================================================

# ---------- Skills ----------
# Columns: A=PRN  B=Skill Type(Technical/Soft/Language)  C=Name  D=Domain  E=Proficiency
@router.post("/import/skills", response_model=dict)
async def import_skills(file: UploadFile = File(...), db: Session = Depends(get_db), admin: User = Depends(get_current_admin)):
    from app.models.skill import TechnicalSkill, SoftSkill, Language
    sheet = _open_sheet(await file.read())
    created, skipped, errors = 0, 0, []

    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(c is None for c in row):
            continue
        prn, skill_type, name, domain, proficiency = (list(row) + [None] * 5)[:5]

        student = _student_by_prn(db, prn)
        if not student:
            errors.append(f"Row {row_num}: no student with PRN {prn}")
            continue
        if not name or not str(name).strip():
            errors.append(f"Row {row_num}: missing skill name")
            continue

        st = (str(skill_type).strip().lower() if skill_type else "")
        try:
            if st == "technical":
                prof = int(proficiency) if proficiency is not None else None
                if prof is None or not (1 <= prof <= 5):
                    errors.append(f"Row {row_num}: technical proficiency must be 1-5"); continue
                if not domain or not str(domain).strip():
                    errors.append(f"Row {row_num}: technical skill needs a domain"); continue
                db.add(TechnicalSkill(user_id=student.id, domain=str(domain).strip(), name=str(name).strip(), proficiency=prof))
            elif st == "soft":
                prof = int(proficiency) if proficiency is not None else None
                if prof is not None and not (1 <= prof <= 5):
                    errors.append(f"Row {row_num}: soft proficiency must be 1-5"); continue
                db.add(SoftSkill(user_id=student.id, name=str(name).strip(), proficiency=prof))
            elif st == "language":
                allowed = {"Basic", "Conversational", "Proficient", "Fluent", "Native"}
                pv = str(proficiency).strip().title() if proficiency else ""
                if pv not in allowed:
                    errors.append(f"Row {row_num}: language proficiency must be one of {', '.join(allowed)}"); continue
                db.add(Language(user_id=student.id, name=str(name).strip(), proficiency=pv))
            else:
                errors.append(f"Row {row_num}: skill type must be Technical, Soft, or Language"); continue
            db.flush()
            created += 1
        except (ValueError, TypeError, SQLAlchemyError) as e:
            db.rollback()
            errors.append(f"Row {row_num}: {e}")

    db.commit()
    return {"created": created, "skipped": skipped, "errors": errors}


# ---------- Projects ----------
# Columns: A=PRN B=Name C=Description D=Domain E=TechStack(comma-sep) F=Type G=MentorName H=Status I=StartDate J=EndDate
@router.post("/import/projects", response_model=dict)
async def import_projects(file: UploadFile = File(...), db: Session = Depends(get_db), admin: User = Depends(get_current_admin)):
    from app.models.project import Project
    sheet = _open_sheet(await file.read())
    created, skipped, errors = 0, 0, []
    valid_types = {"College Project", "Personal Project", "Internship Project"}
    valid_status = {"Ongoing", "Completed"}

    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(c is None for c in row):
            continue
        prn, name, desc, domain, tech, ptype, mentor, pstatus, start, end = (list(row) + [None] * 10)[:10]

        student = _student_by_prn(db, prn)
        if not student:
            errors.append(f"Row {row_num}: no student with PRN {prn}"); continue
        if not name or not str(name).strip():
            errors.append(f"Row {row_num}: missing project name"); continue
        ptype = str(ptype).strip() if ptype else ""
        if ptype not in valid_types:
            errors.append(f"Row {row_num}: type must be one of {', '.join(valid_types)}"); continue
        pstatus = str(pstatus).strip() if pstatus else ""
        if pstatus not in valid_status:
            errors.append(f"Row {row_num}: status must be Ongoing or Completed"); continue
        if ptype == "College Project" and (not mentor or not str(mentor).strip()):
            errors.append(f"Row {row_num}: College Project needs a mentor name"); continue

        try:
            sd, ed = _parse_date(start), _parse_date(end)
            if sd and ed and ed < sd:
                errors.append(f"Row {row_num}: end date before start date"); continue
            tech_list = [t.strip() for t in str(tech).split(",")] if tech else []
            tech_list = [t for t in tech_list if t]
            db.add(Project(
                user_id=student.id, name=str(name).strip(),
                description=str(desc).strip() if desc else None,
                domain=str(domain).strip() if domain else None,
                tech_stack=tech_list, type=ptype,
                mentor_name=str(mentor).strip() if mentor else None,
                status=pstatus, start_date=sd, end_date=ed,
            ))
            db.flush()
            created += 1
        except (ValueError, TypeError, SQLAlchemyError) as e:
            db.rollback()
            errors.append(f"Row {row_num}: {e}")

    db.commit()
    return {"created": created, "skipped": skipped, "errors": errors}


# ---------- Achievements ----------
# Columns: A=PRN B=Title C=Description D=Category E=Type F=Level G=Date
@router.post("/import/achievements", response_model=dict)
async def import_achievements(file: UploadFile = File(...), db: Session = Depends(get_db), admin: User = Depends(get_current_admin)):
    from app.models.achievement import Achievement
    sheet = _open_sheet(await file.read())
    created, skipped, errors = 0, 0, []
    valid_cat = {"Academic", "Co-curricular", "Sports", "Technical", "Cultural", "Other"}
    valid_type = {"Competition", "Hackathon", "Award", "Certification", "Publication", "Other"}
    valid_level = {"College", "State", "National", "International"}

    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(c is None for c in row):
            continue
        prn, title, desc, cat, atype, level, adate = (list(row) + [None] * 7)[:7]

        student = _student_by_prn(db, prn)
        if not student:
            errors.append(f"Row {row_num}: no student with PRN {prn}"); continue
        if not title or not str(title).strip():
            errors.append(f"Row {row_num}: missing title"); continue
        cat = str(cat).strip() if cat else ""
        if cat not in valid_cat:
            errors.append(f"Row {row_num}: category must be one of {', '.join(valid_cat)}"); continue
        atype = str(atype).strip() if atype else ""
        if atype not in valid_type:
            errors.append(f"Row {row_num}: type must be one of {', '.join(valid_type)}"); continue
        level = str(level).strip() if level else ""
        if level not in valid_level:
            errors.append(f"Row {row_num}: level must be one of {', '.join(valid_level)}"); continue

        try:
            db.add(Achievement(
                user_id=student.id, title=str(title).strip(),
                description=str(desc).strip() if desc else None,
                category=cat, type=atype, level=level, date=_parse_date(adate),
            ))
            db.flush()
            created += 1
        except (ValueError, TypeError, SQLAlchemyError) as e:
            db.rollback()
            errors.append(f"Row {row_num}: {e}")

    db.commit()
    return {"created": created, "skipped": skipped, "errors": errors}


# ---------- Work Experience ----------
# Columns: A=PRN B=Organisation C=Role D=Type E=StartDate F=EndDate G=Description
@router.post("/import/work-experience", response_model=dict)
async def import_work_experience(file: UploadFile = File(...), db: Session = Depends(get_db), admin: User = Depends(get_current_admin)):
    from app.models.experience import Experience
    sheet = _open_sheet(await file.read())
    created, skipped, errors = 0, 0, []
    valid_type = {"Internship", "Part-time", "Full-time"}

    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(c is None for c in row):
            continue
        prn, org, role, etype, start, end, desc = (list(row) + [None] * 7)[:7]

        student = _student_by_prn(db, prn)
        if not student:
            errors.append(f"Row {row_num}: no student with PRN {prn}"); continue
        if not org or not str(org).strip():
            errors.append(f"Row {row_num}: missing organisation name"); continue
        if not role or not str(role).strip():
            errors.append(f"Row {row_num}: missing role"); continue
        etype = str(etype).strip() if etype else ""
        if etype not in valid_type:
            errors.append(f"Row {row_num}: type must be Internship, Part-time, or Full-time"); continue

        try:
            sd, ed = _parse_date(start), _parse_date(end)
            if sd and ed and ed < sd:
                errors.append(f"Row {row_num}: end date before start date"); continue
            db.add(Experience(
                user_id=student.id, organisation_name=str(org).strip(), role=str(role).strip(),
                type=etype, start_date=sd, end_date=ed,
                description=str(desc).strip() if desc else None,
            ))
            db.flush()
            created += 1
        except (ValueError, TypeError, SQLAlchemyError) as e:
            db.rollback()
            errors.append(f"Row {row_num}: {e}")

    db.commit()
    return {"created": created, "skipped": skipped, "errors": errors}


# ============================================================
#  Full user list for admin UI (name, department, status)
# ============================================================
@router.get("/users-list", response_model=list[AdminUserResponse])
async def list_users_full(
    role: str | None = None,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    """List users with full details (name, department) for admin UI."""
    query = db.query(User)
    if role:
        query = query.filter(User.role == role)
    return query.all()


@router.patch("/cohorts/{cohort_id}")
async def update_cohort_mentor(
    cohort_id: str,
    payload: CohortMentorUpdate,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    """Assign (or clear) a cohort's academic mentor.
    Writes MentorAssignment rows for every student in the cohort so the
    teacher portal reflects the assignment."""
    cohort = db.query(Cohort).filter(Cohort.id == cohort_id).first()
    if cohort is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Cohort not found")

    mentor_id = payload.academic_mentor_id

    # Validate the mentor if one is provided
    if mentor_id:
        mentor = db.query(User).filter(User.id == mentor_id, User.role == "teacher").first()
        if mentor is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Teacher not found")
    else:
        mentor = None

    # Set the cohort's mentor
    cohort.mentor_id = mentor_id

    # Find all students in this cohort (matching year + department)
    students = db.query(User).filter(
        User.role == "student",
        User.academic_year == cohort.academic_year,
        User.department == cohort.department,
    ).all()

    # Update each student's MentorAssignment to reflect the cohort mentor
    for student in students:
        existing = db.query(MentorAssignment).filter(
            MentorAssignment.student_id == student.id
        ).first()
        if mentor_id:
            if existing:
                existing.teacher_id = mentor_id
            else:
                db.add(MentorAssignment(teacher_id=mentor_id, student_id=student.id))
        else:
            # clearing the mentor: remove existing assignment
            if existing:
                db.delete(existing)

    db.commit()
    db.refresh(cohort)

    return {
        "id": cohort.id,
        "academicYear": cohort.academic_year,
        "department": cohort.department,
        "studentCount": len(students),
        "academicMentorId": cohort.mentor_id,
        "academicMentorName": mentor.name if mentor else None,
    }
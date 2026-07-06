from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Response
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from app.core.dependencies import get_current_admin, get_db
from app.models.user import User
from app.models.marks import MarkingScheme, SchemeComponent, MarksEntry
from app.models.academic_structure import Course, Division
from app.schemas.marks import SchemeCreate, SchemeUpdate, SchemeResponse
from app.core.marks_engine import (
    compute_student_cgpa, compute_subject_result, recompute_and_cache_cgpa,
)

router = APIRouter(prefix="/admin/marking-schemes", tags=["marking-schemes"])


def _commit_or_conflict(db: Session, detail: str):
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=detail)


def _load_components(db: Session, scheme_id: str) -> list:
    return db.query(SchemeComponent).filter(
        SchemeComponent.scheme_id == scheme_id
    ).order_by(SchemeComponent.display_order, SchemeComponent.code).all()


def _scheme_with_components(db: Session, scheme: MarkingScheme) -> dict:
    return {
        "id": scheme.id,
        "name": scheme.name,
        "description": scheme.description,
        "components": _load_components(db, scheme.id),
    }


@router.post("", response_model=SchemeResponse, status_code=status.HTTP_201_CREATED)
async def create_scheme(
    payload: SchemeCreate,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    codes = [c.code for c in payload.components]
    if len(codes) != len(set(codes)):
        raise HTTPException(status_code=422, detail="Duplicate component codes in the scheme")

    scheme = MarkingScheme(name=payload.name, description=payload.description)
    db.add(scheme)
    try:
        db.flush()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="A scheme with that name already exists")

    for c in payload.components:
        db.add(SchemeComponent(
            scheme_id=scheme.id, code=c.code, label=c.label,
            max_marks=c.max_marks, min_marks=c.min_marks, display_order=c.display_order,
        ))
    _commit_or_conflict(db, "A scheme with that name already exists")
    db.refresh(scheme)
    return _scheme_with_components(db, scheme)


@router.get("", response_model=list[SchemeResponse])
async def list_schemes(
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    schemes = db.query(MarkingScheme).order_by(MarkingScheme.name).all()
    return [_scheme_with_components(db, s) for s in schemes]


@router.get("/{scheme_id}", response_model=SchemeResponse)
async def get_scheme(
    scheme_id: str,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    scheme = db.query(MarkingScheme).filter(MarkingScheme.id == scheme_id).first()
    if scheme is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Marking scheme not found")
    return _scheme_with_components(db, scheme)


@router.patch("/{scheme_id}", response_model=SchemeResponse)
async def update_scheme(
    scheme_id: str,
    payload: SchemeUpdate,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    scheme = db.query(MarkingScheme).filter(MarkingScheme.id == scheme_id).first()
    if scheme is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Marking scheme not found")

    if payload.name is not None:
        scheme.name = payload.name
    if payload.description is not None:
        scheme.description = payload.description

    if payload.components is not None:
        codes = [c.code for c in payload.components]
        if len(codes) != len(set(codes)):
            raise HTTPException(status_code=422, detail="Duplicate component codes in the scheme")
        used = db.query(MarksEntry).join(
            SchemeComponent, MarksEntry.component_id == SchemeComponent.id
        ).filter(SchemeComponent.scheme_id == scheme_id).first()
        if used:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Cannot change components: marks have already been entered using this scheme.",
            )
        db.query(SchemeComponent).filter(SchemeComponent.scheme_id == scheme_id).delete()
        for c in payload.components:
            db.add(SchemeComponent(
                scheme_id=scheme_id, code=c.code, label=c.label,
                max_marks=c.max_marks, min_marks=c.min_marks, display_order=c.display_order,
            ))

    _commit_or_conflict(db, "A scheme with that name already exists")
    db.refresh(scheme)
    return _scheme_with_components(db, scheme)


@router.delete("/{scheme_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_scheme(
    scheme_id: str,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    scheme = db.query(MarkingScheme).filter(MarkingScheme.id == scheme_id).first()
    if scheme is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Marking scheme not found")

    if db.query(Course).filter(Course.marking_scheme_id == scheme_id).first():
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Cannot delete a scheme that is assigned to courses. Reassign those courses first.",
        )
    db.query(SchemeComponent).filter(SchemeComponent.scheme_id == scheme_id).delete()
    db.delete(scheme)
    db.commit()

@router.post("/marks/upload/{course_id}", response_model=dict, tags=["marks"])
async def upload_marks(
    course_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    """Upload marks for ONE course. Header row must contain component CODES
    (CCA-TH, ETE-TH, etc.). Use 'AB' in a cell to mark a student absent.
    Re-uploading updates existing marks."""
    course = db.query(Course).filter(Course.id == course_id).first()
    if course is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Course not found")
    if not course.marking_scheme_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                            detail="This course has no marking scheme assigned")

    components = db.query(SchemeComponent).filter(
        SchemeComponent.scheme_id == course.marking_scheme_id
    ).all()
    if not components:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                            detail="The course's marking scheme has no components")
    code_to_component = {c.code.strip().upper(): c for c in components}

    try:
        wb = load_workbook(BytesIO(await file.read()))
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid Excel file")
    sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Empty spreadsheet")

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    try:
        prn_idx = next(i for i, h in enumerate(header) if h.upper() == "PRN")
    except StopIteration:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No 'PRN' column found")

    col_to_component = {}
    for i, h in enumerate(header):
        comp = code_to_component.get(h.upper())
        if comp:
            col_to_component[i] = comp
    if not col_to_component:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"No component columns found. Expected any of: {', '.join(code_to_component.keys())}",
        )

    updated = 0
    created = 0
    errors = []
    touched_students = set()

    for row_num, row in enumerate(rows[1:], start=2):
        if row is None or all(c is None for c in row):
            continue
        prn = row[prn_idx] if prn_idx < len(row) else None
        if not prn:
            errors.append(f"Row {row_num}: missing PRN")
            continue
        prn = str(prn).strip()
        student = db.query(User).filter(User.prn == prn, User.role == "student").first()
        if not student:
            errors.append(f"Row {row_num}: no student with PRN {prn}")
            continue

        for col_idx, comp in col_to_component.items():
            raw = row[col_idx] if col_idx < len(row) else None

            # blank cell -> skip (no entry for this component)
            if raw is None or (isinstance(raw, str) and raw.strip() == ""):
                continue

            # "AB" / "A" / "ABSENT" -> store as None (absent), not a number
            if isinstance(raw, str) and raw.strip().upper() in ("AB", "A", "ABSENT"):
                val = None
            else:
                try:
                    val = float(raw)
                except (ValueError, TypeError):
                    errors.append(f"Row {row_num}: '{comp.code}' value '{raw}' is not a number (use a number or 'AB' for absent)")
                    continue
                if val < 0 or val > comp.max_marks:
                    errors.append(f"Row {row_num}: '{comp.code}' = {val} out of range 0..{comp.max_marks}")
                    continue

            existing = db.query(MarksEntry).filter(
                MarksEntry.student_id == student.id,
                MarksEntry.course_id == course.id,
                MarksEntry.component_id == comp.id,
            ).first()
            if existing:
                existing.obtained_marks = val
                updated += 1
            else:
                db.add(MarksEntry(student_id=student.id, course_id=course.id,
                                  component_id=comp.id, obtained_marks=val))
                created += 1
            touched_students.add(student.id)

    db.commit()

    for sid in touched_students:
        recompute_and_cache_cgpa(db, sid)

    return {
        "course": course.course_code,
        "created": created,
        "updated": updated,
        "studentsAffected": len(touched_students),
        "errors": errors,
    }


@router.get("/marks/student/{student_id}", response_model=dict, tags=["marks"])
async def get_student_marks(
    student_id: str,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    """Full computed record: every semester, subject breakdown, SGPA, and CGPA."""
    student = db.query(User).filter(User.id == student_id, User.role == "student").first()
    if student is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Student not found")
    result = compute_student_cgpa(db, student_id)
    return {
        "studentId": student.id,
        "prn": student.prn,
        "name": student.name,
        "cgpa": result["cgpa"],
        "totalCredits": result["total_credits"],
        "semesters": result["semesters"],
    }

@router.get("/marks/template/{course_id}", tags=["marks"])
async def download_marks_template(
    course_id: str,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    """Generate a ready-to-fill Excel: PRN, Name, component columns (max in header),
    one row per enrolled student (same branch + semester as the course)."""
    course = db.query(Course).filter(Course.id == course_id).first()
    if course is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Course not found")
    if not course.marking_scheme_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                            detail="This course has no marking scheme assigned")

    components = db.query(SchemeComponent).filter(
        SchemeComponent.scheme_id == course.marking_scheme_id
    ).order_by(SchemeComponent.display_order, SchemeComponent.code).all()
    if not components:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                            detail="The course's marking scheme has no components")

    division_ids = [d.id for d in db.query(Division).filter(
        Division.branch_id == course.branch_id
    ).all()]
    students = []
    if division_ids:
        students = db.query(User).filter(
            User.role == "student",
            User.division_id.in_(division_ids),
            User.current_semester == course.semester,
        ).order_by(User.prn).all()

    wb = Workbook()
    ws = wb.active
    ws.title = (course.course_code or "Marks")[:31]

    headers = ["PRN", "Name"] + [
        f"{c.code} ({int(c.max_marks) if c.max_marks == int(c.max_marks) else c.max_marks})"
        for c in components
    ]
    ws.append(headers)

    fill = PatternFill(start_color="1F2A44", end_color="1F2A44", fill_type="solid")
    for i in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=i)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    for s in students:
        ws.append([s.prn, s.name] + ["" for _ in components])

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 24
    for idx in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 14

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"marks_template_{course.course_code}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/marks/course/{course_id}", response_model=dict, tags=["marks"])
async def list_course_marks(
    course_id: str,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    """List every student's computed result for ONE course — a class marksheet view."""
    from app.core.marks_engine import compute_subject_result

    course = db.query(Course).filter(Course.id == course_id).first()
    if course is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Course not found")

    student_ids = [r.student_id for r in db.query(MarksEntry.student_id).filter(
        MarksEntry.course_id == course_id
    ).distinct().all()]

    students = db.query(User).filter(User.id.in_(student_ids)).order_by(User.prn).all() if student_ids else []

    results = []
    passed_count = 0
    failed_count = 0
    for s in students:
        res = compute_subject_result(db, s.id, course)
        if not res:
            continue
        if res["passed"]:
            passed_count += 1
        else:
            failed_count += 1
        results.append({
            "studentId": s.id,
            "prn": s.prn,
            "name": s.name,
            "obtainedTotal": res["obtained_total"],
            "maxTotal": res["max_total"],
            "percentage": res["percentage"],
            "grade": res["grade"],
            "gradePoints": res["grade_points"],
            "passed": res["passed"],
            "hasAbsent": res["has_absent"],
        })

    return {
        "courseId": course.id,
        "courseCode": course.course_code,
        "courseName": course.course_name,
        "credits": course.credits,
        "studentsWithMarks": len(results),
        "passed": passed_count,
        "failed": failed_count,
        "students": results,
    }


@router.delete("/marks/course/{course_id}", response_model=dict, tags=["marks"])
async def delete_course_marks(
    course_id: str,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    """Delete ALL marks for a course (e.g. to redo a bad upload)."""
    course = db.query(Course).filter(Course.id == course_id).first()
    if course is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Course not found")

    affected = [r.student_id for r in db.query(MarksEntry.student_id).filter(
        MarksEntry.course_id == course_id
    ).distinct().all()]

    deleted = db.query(MarksEntry).filter(MarksEntry.course_id == course_id).delete()
    db.commit()

    for sid in affected:
        recompute_and_cache_cgpa(db, sid)

    return {
        "courseCode": course.course_code,
        "deletedEntries": deleted,
        "studentsAffected": len(affected),
    }